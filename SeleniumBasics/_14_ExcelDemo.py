import openpyxl

book = openpyxl.load_workbook("H:\\Pycharm Projects\\SeleniumPythonProject\\SeleniumBasics\\TestData.xlsx")
sheet = book.active

#get cell
cell = sheet.cell(row=1, column=1)
print(cell.value)

#write to excel cell - and then save the worbook
# sheet.cell(row=5, column=1).value = "NewTestCase"
# sheet['B5']="NewName"
# book.save("H:\\Pycharm Projects\\SeleniumPythonProject\\SeleniumBasics\\TestData.xlsx")

#get row count and column count
print(sheet.max_row)
print(sheet.max_column)

#get value
print(sheet['B2'].value)

# iterate through sheet
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value, end="\t")
    print()

#store data in a dictionary for TC2
dict = {}
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TC2":
        for j in range(2,sheet.max_column+1):
            dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value

print(dict)





